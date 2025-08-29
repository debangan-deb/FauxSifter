from flask import Flask, request, send_file, jsonify
from flask_cors import CORS
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from pathlib import Path
import re, requests, io, time, json, pandas as pd, joblib
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from PIL import Image as PILImage, ImageFile
ImageFile.LOAD_TRUNCATED_IMAGES=True

app=Flask(__name__); CORS(app)
BASE=Path(__file__).resolve().parent
EXCEL="predicted_reviews.xlsx"
rurl=lambda a,n=1:f"https://www.amazon.in/product-reviews/{a}/?reviewerType=all_reviews&pageNumber={n}"
dpurl=lambda a:f"https://www.amazon.in/dp/{a}"

@app.get("/")
def health(): return "ok",200

def safe_goto(pg,url,wait="domcontentloaded",timeout=30000,retries=1):
    for i in range(retries+1):
        try: return pg.goto(url,wait_until=wait,timeout=timeout)
        except PWTimeout:
            if i<retries: time.sleep(1.2)
            else: raise

def extract_title(pg):
    for sel in ["span#productTitle","h1#title span","#title span","[data-feature-name='title'] span#productTitle"]:
        try:
            t=(pg.locator(sel).first.text_content(timeout=3000) or "").strip()
            if t and not re.search(r"\b\d+\s*star\b|\b\d+\s*%\b|\bpercent\b",t.lower()): return t
        except: pass
    try:
        t=(pg.eval_on_selector("meta[property='og:title']","e=>e.getAttribute('content')") or "").strip()
        if t: return t
    except: pass
    try:
        t=(pg.title() or ""); t=re.sub(r"\s*\|\s*Amazon\.in\s*$","",t); t=re.sub(r"^Amazon\.in:\s*Buy\s+","",t)
        if t.strip(): return t.strip()
    except: pass

def colpx(w): return int((w*7+5) if w>=1 else w*12)
def rowpx(hpt): return int(hpt*4/3)

@app.post("/predict")
def predict():
    m=re.search(r"/dp/([A-Z0-9]{10})",(request.get_json() or {}).get("url","")); 
    if not m: return "",404
    asin=m.group(1); rows,seen,title,prod_img=[],set(),"UNKNOWN",None
    try:
        with sync_playwright() as p:
            br=p.chromium.launch(headless=True,args=["--disable-dev-shm-usage","--no-sandbox"])
            st=BASE/"amazon_auth.json"
            ua=("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36")
            kw=dict(viewport={"width":1280,"height":1800},locale="en-IN",user_agent=ua)
            ctx=br.new_context(storage_state=str(st),**kw) if st.exists() else br.new_context(**kw)
            ctx.route("**/*",lambda r:r.abort() if r.request.resource_type in ("image","media","font") else r.continue_())
            pg=ctx.new_page(); pg.set_default_timeout(15000); pg.set_default_navigation_timeout(45000)
            safe_goto(pg,rurl(asin,1),timeout=35000,retries=1); pg.wait_for_selector('[data-hook="review"]',timeout=15000)
            def scrape(p):
                data=p.eval_on_selector_all('[data-hook="review"]',"""
                  els=>els.map(e=>{const g=s=>e.querySelector(s),t=n=>n?(n.textContent||'').trim():'',a=s=>Array.from(e.querySelectorAll(s)).map(x=>x.textContent.trim()).join(' ').trim();
                  const rEl=g('[data-hook="review-star-rating"] .a-icon-alt')||g('.a-icon-alt')||Array.from(e.querySelectorAll('[aria-label]')).find(x=>(x.getAttribute('aria-label')||'').toLowerCase().includes('out of 5 stars'));
                  const rStr=rEl?(rEl.textContent||rEl.getAttribute('aria-label')||'').trim():'',rating=(rStr.match(/([0-9]+(?:\\.[0-9]+)?)/)||['',''])[1]||'';
                  return{id:e.id||'',tt:a('[data-hook="review-title"] span'),bd:a('[data-hook="review-body"] span'),nm:t(g('.a-profile-name')),dt:t(g('[data-hook="review-date"]')),rt:rating};})
                """)
                for r in data:
                    if not r["id"] or r["id"] in seen: continue
                    tt=(r["tt"] or "").strip(); bd=(r["bd"] or "").strip(); used=bd or tt
                    if not used: continue
                    rows.append([None,tt,bd,used,"body" if bd else "title",(r["nm"] or "").strip(),(r["dt"] or "").strip(),(r["rt"] or "").strip()]); seen.add(r["id"])
            scrape(pg); c=1
            while c<120:
                last=pg.locator("ul.a-pagination li.a-last")
                if not last.count(): break
                try: dis=last.evaluate("el=>el.classList.contains('a-disabled')||el.getAttribute('aria-disabled')==='true'")
                except: dis=True
                if dis: break
                prev=set(pg.eval_on_selector_all('[data-hook="review"]',"els=>els.map(e=>e.id||'')"))
                pg.evaluate("window.scrollTo(0, document.body.scrollHeight)"); last.locator("a").click()
                try: pg.wait_for_function("""prev=>{const ids=[...document.querySelectorAll('[data-hook="review"]')].map(e=>e.id||'');if(!ids.length)return false;const s=new Set(prev||[]);if(ids.length!==s.size)return true;return ids.some(id=>!s.has(id));}""",arg=list(prev),timeout=9000)
                except: time.sleep(0.4)
                try: pg.wait_for_selector('[data-hook="review"]',timeout=7000)
                except: break
                c+=1; scrape(pg)
            safe_goto(pg,dpurl(asin),timeout=35000,retries=1)
            t=extract_title(pg); title=t or title
            def big_img(p):
                for s in ["img#landingImage","div#imgTagWrapperId img","img[data-old-hires]","img[data-a-dynamic-image]"]:
                    try:
                        el=p.locator(s).first
                        if not el.count(): continue
                        u=el.get_attribute("src",timeout=1500) or el.get_attribute("data-old-hires",timeout=1500)
                        if u and u.startswith("http"): return u
                        dai=el.get_attribute("data-a-dynamic-image",timeout=1500)
                        if dai:
                            d=json.loads(dai); return sorted(d.keys(),key=lambda k:-(d[k][0]*d[k][1]))[0]
                    except: pass
            prod_img=big_img(pg); br.close()
    except PWTimeout: return jsonify(error="timeout"),504
    except Exception as e: print("playwright:",repr(e)); return jsonify(error="scrape failed"),500

    img_buf=img_size=None
    if prod_img:
        try:
            r=requests.get(prod_img,headers={'User-Agent':"Mozilla/5.0",'Referer':dpurl(asin)},timeout=15); r.raise_for_status()
            im=PILImage.open(io.BytesIO(r.content))
            if im.mode in ("P","RGBA","LA"): im=im.convert("RGB")
            img_size=im.size
            img_buf=io.BytesIO(); im.save(img_buf,format="PNG"); img_buf.seek(0)
        except Exception as e: print("img:",e)

    if not rows: return "",404

    df=pd.DataFrame(rows,columns=["SL NO","Review Title","Review Text","Used Text","Text Source","Reviewer Name","Review Date","Rating"]); df["SL NO"]=range(1,len(df)+1)
    out=io.BytesIO()
    with pd.ExcelWriter(out,engine="xlsxwriter") as w:
        df.to_excel(w,index=False,header=False,startrow=15,startcol=0,sheet_name="Sheet1")
        sh=w.sheets['Sheet1']; wb=w.book
        h=wb.add_format({'bold':True,'font_name':'Arial','font_size':14,'align':'center','valign':'vcenter'})
        n=wb.add_format({'font_name':'Arial','font_size':14,'align':'center','valign':'vcenter'})
        sh.set_column('A:C',30); sh.set_column('D:I',32); sh.set_column('J:J',20); [sh.set_row(r,28) for r in range(0,11)]
        W=colpx(30)*3; H=sum(rowpx(28) for _ in range(10))
        if img_buf and img_size:
            iw,ih=img_size
            pad=8
            s=min((W-2*pad)/iw,(H-2*pad)/ih)
            sw,shp=int(iw*s),int(ih*s)
            xo=max(int((W-sw)/2),0); yo=max(int((H-shp)/2),0)
            sh.insert_image('A1','product.png',{'image_data':img_buf,'x_scale':s,'y_scale':s,'x_offset':xo,'y_offset':yo})
        sh.merge_range('A1:C10','',n); sh.merge_range('D1:I10',(title or "UNKNOWN").upper(),h)
        sh.write_row('A14',["SL NO","REVIEW TITLE","REVIEW TEXT","USED TEXT","TEXT SOURCE","REVIEWER","REVIEW DATE","RATING"],h); sh.write('I14','LABELS',h)
    out.seek(0)

    wb2=load_workbook(out); ws=wb2["Sheet1"]; al=Alignment(horizontal='center',vertical='center')
    for r in ws.iter_rows(min_row=1,max_row=ws.max_row,min_col=1,max_col=ws.max_column):
        for c in r: c.alignment=al; c.font=Font(name='Arial',size=14,bold=(c.row==14 or (1<=c.row<=10 and 4<=c.column<=9)))
    ws["D1"].value=(title or "UNKNOWN").upper(); ws["D1"].font=Font(name='Arial',size=14,bold=True); ws["D1"].alignment=al

    model=joblib.load(BASE/"svm_model.pkl"); texts,idx=[],[]
    for r in range(16,ws.max_row+1):
        v=ws.cell(row=r,column=4).value
        if v and str(v).strip(): texts.append(str(v)); idx.append(r)
    if not texts:
        bio=io.BytesIO(); wb2.save(bio); wb2.close(); bio.seek(0); return "",404

    preds=model.predict(pd.Series(texts))
    for r,p in zip(idx,preds):
        cell=ws.cell(row=r,column=9,value="REAL" if p else "FAKE"); cell.font=Font(name='Arial',size=14,bold=not p)
        if not p: cell.fill=PatternFill(start_color="FFFF00",end_color="FFFF00",fill_type="solid")

    final=io.BytesIO(); wb2.save(final); wb2.close(); final.seek(0)
    return send_file(final,as_attachment=True,download_name=EXCEL,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

if __name__=="__main__": app.run()
